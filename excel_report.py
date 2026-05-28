"""
excel_report.py — Excel 总表生成模块
生成 output/朱哥短线雷达_交易复盘总表.xlsx。失败静默不中断主流程。
"""
import logging
import math
from datetime import date as _date, timedelta
from pathlib import Path
from typing import Optional

import pandas as pd

logger = logging.getLogger(__name__)

BASE_DIR    = Path(__file__).parent
OUTPUT_DIR  = BASE_DIR / "output"
CSV_PATH    = BASE_DIR / "output" / "trade_review.csv"
EXCEL_PATH  = OUTPUT_DIR / "朱哥短线雷达_交易复盘总表.xlsx"

# ── colour palette ────────────────────────────────────────────────────────────
_C_HEADER  = "2E75B6"   # header bg (blue)
_C_HDR_TXT = "FFFFFF"   # header text (white)
_C_SECTION = "D6E4F0"   # section label bg (light blue)
_C_ALT     = "F2F8FD"   # alternating row bg (very light blue)
_C_TITLE   = "1F3864"   # title bg (dark navy)

# ── translation tables ────────────────────────────────────────────────────────
_MODE_CN = {"full": "全A模式", "theme_auto": "主题龙头模式"}

_NOTES_CN = {
    # V1.3
    "market_sentiment_below_5":        "大盘情绪不足5分",
    "open_change_too_high":            "开盘涨幅超过4%，高开过多",
    "open_change_too_low":             "开盘跌幅超过1%，开盘偏弱",
    "price_below_open":                "9:36价格低于开盘价，承接不足",
    "price_below_ma5":                 "9:36价格低于5日线，短线走弱",
    "unable_to_buy_limit_up":          "一字涨停买不进",
    "possible_limit_up_unable_to_buy": "疑似涨停买不进",
    # V1.4 新增（主因）
    "theme_strength_too_low":          "主题强度不足，暂不买入",
    "full_score_not_strong_enough":    "全A模式分数或人气技术不够强，只观察不买入",
    "open_change_too_low_hard":        "开盘跌幅超过3%，明显弱开，直接放弃",
    # V1.4 新增（辅助）
    "open_change_weak_watch":          "低开超过1%，开盘偏弱，但不单独否决",
}

# _FIELD_NOTES: (name, description) tuples.
# name == "" → rendered as a section header row using description text.
_FIELD_NOTES = [
    # ── 选股评分 ──────────────────────────────────────────────────────────────
    ("", "▌ 选股评分字段"),
    ("总分",             "系统综合打分，最高约100分，越高越优先推荐。总分高不代表买入，还需通过9:36确认。"),
    ("人气分",           "衡量资金关注度，主要看成交量比率、换手率、量能变化。高分说明资金主动参与。"),
    ("技术分",           "衡量短线趋势强度，主要看均线结构、MACD、量价配合。高分说明趋势向上、形态健康。"),
    ("空间分",           "衡量上涨空间，主要看距近期高点距离和流通盘大小。高分说明没有严重透支，还有上涨余地。"),
    ("风险分",           "风险扣分项，有ST风险、连续大涨、量价背离等情况会扣分。分数越低说明系统认为风险越小。"),
    ("大盘情绪分",       "当日大盘情绪（0-10分）。≥5才允许模拟买入。V1.3版本仅用上证指数+成交额，不含涨停家数、炸板率等专业指标。"),
    # ── 主题龙头模式 ─────────────────────────────────────────────────────────
    ("", "▌ 主题龙头模式专属字段（全A模式此处为空）"),
    ("主题强度",         "该主题今日板块综合活跃度（0-100），综合看板块涨幅、成交额、上涨家数占比。越高说明这个方向今天越活跃。"),
    ("主题加分",         "因所属强势主题获得的额外加分，最高20分。主题强度越高，加分越多。"),
    ("主题模式分",       "主题龙头模式排名分 = 系统总分 × 0.8 + 主题加分。乘以0.8是为给主题加分留出空间，该系数是实验值。"),
    # ── 买入价格 ─────────────────────────────────────────────────────────────
    ("", "▌ 模拟买入价格字段"),
    ("模拟买入价",       "满足条件时记录的买入触发价，取9:36实时价格。仅在「是否模拟买入=是」时有值。"),
    ("滑点后买入价",     "模拟买入价 × 1.001（+0.1%滑点），模拟真实报价延迟和买卖价差。所有收益率均以此价格为基准计算。"),
    ("止损价",           "滑点后买入价 × 0.97，亏损达到 -3% 时的止损线。跌破此价按止损价出局。"),
    # ── T+1 复盘结果 ─────────────────────────────────────────────────────────
    ("", "▌ T+1 复盘结果字段"),
    ("次日最高收益",     "（次日最高价 − 滑点后买入价）/ 滑点后买入价。代表买入后给过的最大获利机会，不代表能真的卖到最高。"),
    ("最大回撤",         "（次日最低价 − 滑点后买入价）/ 滑点后买入价。负数，绝对值越大说明最大浮亏越深。"),
    ("模拟交易收益",     "最终模拟盈亏。触发止损按止损价结算；隔夜低开过止损按次日开盘价结算；否则按次日收盘价结算。"),
    ("风险调整后是否成功", "冲高≥3% 且未触发止损，两个条件同时满足才算成功。这是最严格也最实用的单笔成功标准。"),
    ("路径是否不确定",   "同日既冲高≥3%又触发止损，只看日线数据无法判断是先涨后跌还是先跌后涨，需人工查看分时图确认。"),
    # ── 统计汇总指标 ─────────────────────────────────────────────────────────
    ("", "▌ 统计汇总指标（模式统计 sheet）"),
    ("买入触发率",       "有效推荐样本中，满足9:36买入条件的比例。太低（<20%）说明条件过严；太高（>70%）说明条件过松。合理区间约30-60%。"),
    ("风险调整后成功率", "有效买入样本中，冲高≥3%且未先触止损的比例。这是最核心的选股质量指标。"),
    ("盈亏比",           "平均盈利 ÷ 平均亏损绝对值。大于1说明赚的比亏的多；参考值：>1.5为优秀。"),
    ("止损率",           "触发-3%止损线的比例。越低越好；止损率高说明选股质量或市场环境差。"),
    # ── 有效样本 ─────────────────────────────────────────────────────────────
    ("", "▌ 有效样本定义"),
    ("有效推荐样本",     "已完成9:36检查的推荐票（open_price字段有值）。是买入触发率的分母。"),
    ("有效买入样本",     "buy_signal=true 且已有T+1复盘数据。参与胜率、盈亏比、止损率等核心统计。"),
    ("未买入观察样本",   "未触发买入，T+1后观察「如果买了会怎样」，验证选股质量。不计入正式胜率统计。"),
    # ── V1.4 买入标准 ────────────────────────────────────────────────────────
    ("", "▌ V1.4 买入标准 = 逻辑 + 资金 + 买点 + 风险（自 20260527 启用）"),
    ("有效推荐预闸·theme_auto", "theme_strength ≥ 50 才允许进入9:36买入确认；不足只观察。原因码：theme_strength_too_low"),
    ("有效推荐预闸·full",       "总分 ≥ 78 且 人气分 ≥ 22 且 技术分 ≥ 20 才允许进入9:36买入确认；任一不达标只观察。原因码：full_score_not_strong_enough"),
    ("开盘涨幅·硬否决（低）",   "开盘涨幅 < -3% 直接放弃。原因码：open_change_too_low_hard"),
    ("开盘涨幅·辅助提示",       "[-3%, -1%) 视为低开观察，不单独否决；若 9:36 仍站上开盘价和5日线，允许模拟买入。原因码：open_change_weak_watch"),
    ("开盘涨幅·硬否决（高）",   "开盘涨幅 > +4% 直接放弃，避免追高。原因码：open_change_too_high"),
    ("9:36 四条共有标准",       "① 大盘情绪 ≥5  ② 9:36价 ≥ 开盘价  ③ 9:36价 ≥ 5日均线  ④ 非一字涨停"),
    ("买入四因·逻辑",           "属于强主题（theme_strength ≥50）或全A高分强势票（总分 ≥78）"),
    ("买入四因·资金",           "人气分 ≥22，配合昨日成交额体现资金参与度"),
    ("买入四因·买点",           "9:36价同时站上开盘价和5日均线，承接和趋势均确立"),
    ("买入四因·风险",           "开盘涨幅落在 [-3%, +4%] 内，非一字涨停可成交"),
    # ── V1.4 二次确认观察 ──────────────────────────────────────────────────
    ("", "▌ 10:00 二次确认观察（V1.4 实验性观察项 — 仅记录，不计入正式收益）"),
    ("适用对象",                "当日 9:36 检查已完成且 buy_signal_0935=false 的票"),
    ("可二次观察的失败原因",     "price_below_open、price_below_ma5、open_change_weak_watch（其中之一即可，且不含黑名单原因）"),
    ("不可二次观察的失败原因",   "market_sentiment_below_5、theme_strength_too_low、full_score_not_strong_enough、open_change_too_low_hard、open_change_too_high、unable_to_buy_limit_up"),
    ("二次观察通过条件",         "10:00价 ≥ 开盘价 且 10:00价 ≥ 5日均线 且 10:00价 > 9:36价 且 非一字涨停"),
    ("10点价格",                 "10:00 二次观察时的实时价（写入 price_1000 字段）"),
    ("二次观察价",               "通过时的观察买入价；未通过为空。不参与任何收益统计"),
    ("二次观察是否通过",         "true=观察通过（仅记录）；false=未通过；空=今日尚未运行二次观察"),
    ("二次观察结果",             "通过则显示 passed；未通过则按失败码记录原因，可同时多个用';'分隔"),
    ("二次观察时间",             "二次观察实际运行的本地时间（HH:MM:SS）"),
    ("与正式买入的关系",         "二次观察通过 ≠ 模拟买入，不写 buy_signal_0935 / buy_price / 止损价，不进入 T+1 复盘。用于月底分析 9:36 没买但 10:00 走强的票后续表现"),
]


# ── data helpers ──────────────────────────────────────────────────────────────

def _v(x) -> str:
    s = str(x).strip()
    return "" if s in ("nan", "None") else s

def _empty(x) -> bool:
    return _v(x) == ""

def _f(x) -> Optional[float]:
    try:
        v = _v(x)
        if not v:
            return None
        f = float(v)
        return None if math.isnan(f) else f
    except Exception:
        return None

def _gb(x) -> Optional[bool]:
    v = _v(x).lower()
    if v in ("true",  "1"): return True
    if v in ("false", "0"): return False
    return None

def _mode_cn(x) -> str:
    return _MODE_CN.get(_v(x), _v(x) or "未记录")

def _bool_cn(x) -> str:
    v = _v(x).lower()
    if v == "true":  return "是"
    if v == "false": return "否"
    return "未记录"

def _notes_cn(x) -> str:
    raw = _v(x)
    if not raw:
        return ""
    parts = [_NOTES_CN.get(p.strip(), p.strip()) for p in raw.split(";") if p.strip()]
    return "；".join(parts)

def _date_fmt(x) -> str:
    s = _v(x)
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    return s

def _or_unrec(x, fn=None):
    if _empty(x):
        return "未记录"
    return fn(x) if fn else _v(x)

def _gain_str(v: Optional[float]) -> str:
    if v is None:
        return "N/A"
    sign = "+" if v > 0 else ""
    return f"{sign}{v*100:.2f}%"

def _pct_str(v: Optional[float]) -> str:
    if v is None:
        return "N/A"
    return f"{v*100:.1f}%"


# ── stats computation (mirrors trade_review._compute_group_stats) ─────────────

def _compute_stats(df_group: pd.DataFrame) -> dict:
    total = len(df_group)

    valid_mask = df_group.apply(
        lambda r: _f(r.get("open_price")) is not None
               and _f(r.get("price_0935")) is not None,
        axis=1,
    )
    valid_df     = df_group[valid_mask]
    n_valid      = len(valid_df)
    triggered_df = valid_df[valid_df.apply(
        lambda r: _gb(r.get("required_conditions_passed")) is True, axis=1
    )]
    n_triggered  = len(triggered_df)
    bsr          = n_triggered / n_valid if n_valid > 0 else None

    def _is_traded(r):
        return (
            _gb(r.get("buy_signal_0935")) is True
            and _gb(r.get("unable_to_buy")) is not True
            and _f(r.get("buy_price")) is not None
            and _f(r.get("t1_close")) is not None
        )
    traded_df = df_group[df_group.apply(_is_traded, axis=1)]
    n_traded  = len(traded_df)

    empty = dict(
        total=total, n_valid=n_valid, n_triggered=n_triggered,
        bsr=bsr, n_traded=0,
        risk_rate=None, active_rate=None, surge_rate=None,
        close_rate=None, stop_rate=None,
        avg_gain=None, avg_loss=None, wl_ratio=None,
    )
    if n_traded == 0:
        return empty

    def _bools(col):
        return [v for v in traded_df[col].apply(_gb) if v is not None] if col in traded_df.columns else []

    def _floats(col):
        return [v for v in traded_df[col].apply(_f) if v is not None] if col in traded_df.columns else []

    def _rate(bs):
        return sum(bs) / len(bs) if bs else None

    risk_rate   = _rate(_bools("risk_adjusted_success"))
    active_rate = _rate(_bools("is_active_success"))
    surge_rate  = _rate(_bools("is_strong_surge"))
    close_rate  = _rate(_bools("is_close_success"))
    stop_rate   = _rate(_bools("stop_loss_triggered"))

    returns  = _floats("simulated_trade_return")
    gains    = [r for r in returns if r > 0]
    losses   = [r for r in returns if r <= 0]
    avg_gain = sum(gains)  / len(gains)  if gains  else None
    avg_loss = sum(losses) / len(losses) if losses else None
    wl_ratio = abs(avg_gain / avg_loss)  if avg_gain and avg_loss else None

    return dict(
        total=total, n_valid=n_valid, n_triggered=n_triggered,
        bsr=bsr, n_traded=n_traded,
        risk_rate=risk_rate, active_rate=active_rate,
        surge_rate=surge_rate, close_rate=close_rate, stop_rate=stop_rate,
        avg_gain=avg_gain, avg_loss=avg_loss, wl_ratio=wl_ratio,
    )


# ── openpyxl helpers ──────────────────────────────────────────────────────────

def _styles():
    """Return commonly used openpyxl style objects."""
    from openpyxl.styles import Font, PatternFill, Alignment
    hdr_fill   = PatternFill(start_color=_C_HEADER,  end_color=_C_HEADER,  fill_type="solid")
    hdr_font   = Font(bold=True, color=_C_HDR_TXT, size=10)
    sec_fill   = PatternFill(start_color=_C_SECTION, end_color=_C_SECTION, fill_type="solid")
    sec_font   = Font(bold=True, color="1F3864",    size=10)
    alt_fill   = PatternFill(start_color=_C_ALT,    end_color=_C_ALT,    fill_type="solid")
    title_fill = PatternFill(start_color=_C_TITLE,  end_color=_C_TITLE,  fill_type="solid")
    title_font = Font(bold=True, color="FFFFFF", size=12)
    label_font = Font(bold=True, size=10)
    wrap_left  = Alignment(horizontal="left",  vertical="center", wrap_text=True)
    center_va  = Alignment(horizontal="center",vertical="center")
    right_va   = Alignment(horizontal="right", vertical="center")
    return (hdr_fill, hdr_font, sec_fill, sec_font, alt_fill,
            title_fill, title_font, label_font, wrap_left, center_va, right_va)


def _auto_width(ws, max_col_width: int = 50) -> None:
    from openpyxl.utils import get_column_letter
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = 0
        for cell in col:
            if cell.value is not None:
                w = sum(2 if ord(c) > 127 else 1 for c in str(cell.value))
                max_len = max(max_len, w)
        ws.column_dimensions[col_letter].width = min(max_len + 2, max_col_width)


def _write_table(ws, headers: list, rows: list, pct_headers: set,
                 start_row: int) -> int:
    """
    Write a bold header row + data rows to ws.
    pct_headers: column names whose float values are stored as decimals (0.03 → 3%).
    Returns the next empty row index after the table.
    """
    (hdr_fill, hdr_font, _, _, alt_fill,
     _, _, _, wrap_left, center_va, right_va) = _styles()
    from openpyxl.styles import Alignment

    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=start_row, column=col_idx, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center_va

    for r_off, row_dict in enumerate(rows, 1):
        row_idx = start_row + r_off
        use_alt = r_off % 2 == 0
        for col_idx, h in enumerate(headers, 1):
            raw = row_dict.get(h)
            if raw is None or str(raw).strip() in ("", "nan", "None"):
                val = "未记录"
            else:
                val = raw

            cell = ws.cell(row=row_idx, column=col_idx, value=val)
            if use_alt and val != "未记录":
                cell.fill = alt_fill

            if h in pct_headers and isinstance(val, float):
                cell.number_format = "0.00%"
                cell.alignment = right_va
            elif isinstance(val, (int, float)):
                cell.alignment = right_va
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center",
                                           wrap_text=len(str(val)) > 20)

    return start_row + len(rows) + 1


def _write_section_label(ws, row_idx: int, text: str, ncols: int = 8) -> int:
    """Write a merged section label row. Returns next row."""
    (_, _, sec_fill, sec_font, _, _, _, _, _, _, _) = _styles()
    from openpyxl.styles import Alignment
    cell = ws.cell(row=row_idx, column=1, value=text)
    cell.font      = sec_font
    cell.fill      = sec_fill
    cell.alignment = Alignment(horizontal="left", vertical="center")
    if ncols > 1:
        ws.merge_cells(start_row=row_idx, start_column=1,
                       end_row=row_idx,   end_column=ncols)
    return row_idx + 1


# ── overview helpers ──────────────────────────────────────────────────────────

def _buy_status(r) -> str:
    if _empty(r.get("open_price", "")):
        return "待确认"
    if _gb(r.get("unable_to_buy", "")) is True:
        return "⚠️涨停无法成交"
    if _gb(r.get("buy_signal_0935", "")) is True:
        bp = _f(r.get("buy_price", ""))
        return f"✅已买入 @{bp:.2f}" if bp else "✅已买入"
    return "❌未买入"


def _review_status(r) -> str:
    if _gb(r.get("buy_signal_0935", "")) is not True:
        return "—"
    if _empty(r.get("t1_close", "")):
        return "等待T+1数据"
    sim_ret = _f(r.get("simulated_trade_return", ""))
    if sim_ret is None:
        return "—"
    sign = "+" if sim_ret >= 0 else ""
    stop_mark = " ⛔止损" if _gb(r.get("stop_loss_triggered", "")) else ""
    return f"{sign}{sim_ret*100:.2f}%{stop_mark}"


# ── sheet builders ────────────────────────────────────────────────────────────

def _build_overview(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("今日总览")
    (_, _, _, _, _, title_fill, title_font,
     label_font, wrap_left, center_va, _) = _styles()
    from openpyxl.styles import Alignment

    # Determine "today" — use latest report_date in CSV if today has no data
    today_str = _date.today().strftime("%Y%m%d")
    df_today  = df[df["report_date"].apply(_v) == today_str]
    if df_today.empty:
        all_dates = sorted(d for d in df["report_date"].apply(_v).unique() if d)
        if all_dates:
            today_str = all_dates[-1]
            df_today  = df[df["report_date"].apply(_v) == today_str]

    date_fmt  = _date_fmt(today_str)
    df_full   = df_today[df_today["mode"].apply(_v) == "full"].sort_values("rank")
    df_theme  = df_today[df_today["mode"].apply(_v) == "theme_auto"].sort_values("rank")
    sentiment = _v(df_today["market_sentiment"].iloc[0]) if not df_today.empty else "—"
    data_date = _date_fmt(_v(df_today["data_date"].iloc[0])) if not df_today.empty else "—"

    row = 1

    # ── Title ──────────────────────────────────────────────────────────────────
    tc = ws.cell(row=row, column=1, value=f"朱哥短线雷达 — 今日总览  {date_fmt}")
    tc.font = title_font
    tc.fill = title_fill
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row].height = 22
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
    row += 1

    # ── Key info (3 label+value pairs) ────────────────────────────────────────
    for i, (label, value) in enumerate([
        ("报告日期", date_fmt),
        ("数据日期", data_date),
        ("大盘情绪", f"{sentiment}/10" if sentiment not in ("", "—") else "—"),
    ]):
        c = i * 2 + 1
        lc = ws.cell(row=row, column=c,     value=label)
        lc.font = label_font
        vc = ws.cell(row=row, column=c + 1, value=value)
        vc.alignment = Alignment(horizontal="left")
    row += 2

    # ── Full mode ─────────────────────────────────────────────────────────────
    row = _write_section_label(ws, row, "【全A模式推荐（Top 3）】", 8)
    full_hdrs = ["排名", "股票代码", "股票名称", "总分", "开盘状态", "买入状态", "复盘结果"]
    full_rows = []
    for _, r in df_full.iterrows():
        if _empty(r.get("open_price", "")):
            open_st = "待开盘"
        else:
            ocp = _f(r.get("open_change_pct", ""))
            sign = "+" if (ocp or 0) >= 0 else ""
            open_st = f"开盘{sign}{ocp:.2f}%" if ocp is not None else "已开盘"
        full_rows.append({
            "排名": _v(r.get("rank", "")),
            "股票代码": _v(r.get("stock_code", "")),
            "股票名称": _v(r.get("stock_name", "")),
            "总分": _v(r.get("total_score", "")),
            "开盘状态": open_st,
            "买入状态": _buy_status(r),
            "复盘结果": _review_status(r),
        })
    if not full_rows:
        ws.cell(row=row, column=1, value="（今日无全A模式推荐）")
        row += 2
    else:
        row = _write_table(ws, full_hdrs, full_rows, set(), row)
        row += 1

    # ── Theme mode ────────────────────────────────────────────────────────────
    row = _write_section_label(ws, row, "【主题龙头模式推荐（Top 3）】", 8)
    th_hdrs = ["排名", "股票代码", "股票名称", "主题", "主题模式分", "买入状态", "复盘结果"]
    th_rows = []
    for _, r in df_theme.iterrows():
        th_rows.append({
            "排名": _v(r.get("rank", "")),
            "股票代码": _v(r.get("stock_code", "")),
            "股票名称": _v(r.get("stock_name", "")),
            "主题": _v(r.get("theme_name", "")) or "—",
            "主题模式分": _v(r.get("theme_auto_score", "")),
            "买入状态": _buy_status(r),
            "复盘结果": _review_status(r),
        })
    if not th_rows:
        ws.cell(row=row, column=1, value="（今日无主题龙头模式推荐）")
        row += 2
    else:
        row = _write_table(ws, th_hdrs, th_rows, set(), row)
        row += 1

    # ── 已买入 ────────────────────────────────────────────────────────────────
    row = _write_section_label(ws, row, "【今日已触发模拟买入】", 8)
    bought = df_today[
        df_today["buy_signal_0935"].apply(lambda v: _gb(v) is True) &
        ~df_today["unable_to_buy"].apply(lambda v: _gb(v) is True)
    ]
    if bought.empty:
        ws.cell(row=row, column=1, value="今日无模拟买入")
        row += 2
    else:
        b_hdrs = ["模式", "代码", "名称", "买入价", "止损价", "复盘状态"]
        b_rows = []
        for _, r in bought.iterrows():
            bp   = _f(r.get("adjusted_buy_price", "")) or _f(r.get("buy_price", ""))
            stop = _f(r.get("stop_price", ""))
            b_rows.append({
                "模式": _mode_cn(r.get("mode", "")),
                "代码": _v(r.get("stock_code", "")),
                "名称": _v(r.get("stock_name", "")),
                "买入价": f"{bp:.2f}" if bp else "未记录",
                "止损价": f"{stop:.2f}" if stop else "未记录",
                "复盘状态": _review_status(r),
            })
        row = _write_table(ws, b_hdrs, b_rows, set(), row)
        row += 1

    # ── 关键提醒 ──────────────────────────────────────────────────────────────
    row = _write_section_label(ws, row, "【今日关键提醒】", 8)
    alerts: list[str] = []

    for _, r in bought.iterrows():
        stop = _f(r.get("stop_price", ""))
        name = _v(r.get("stock_name", ""))
        if stop:
            alerts.append(f"已买入 {name}，止损价 {stop:.2f}（买入价×0.97），请盯盘")

    for _, r in bought.iterrows():
        if _gb(r.get("stop_loss_triggered", "")) is True:
            sim_ret = _f(r.get("simulated_trade_return", ""))
            name    = _v(r.get("stock_name", ""))
            ret_str = f"{sim_ret*100:.2f}%" if sim_ret is not None else "未知"
            alerts.append(f"⚠️ {name} 已触发止损，模拟亏损 {ret_str}")

    if not alerts:
        if df_today.empty:
            alerts.append("今日暂无推荐数据，等待盘前运行。")
        elif bought.empty:
            sample_reason = ""
            for _, r in df_today.iterrows():
                n = _notes_cn(r.get("notes", ""))
                if n:
                    sample_reason = n
                    break
            if sample_reason:
                alerts.append(f"今日所有推荐均未触发买入，主要原因：{sample_reason}")
            else:
                alerts.append("今日所有推荐均未满足9:36买入条件，或检查尚未运行。")

    from openpyxl.styles import Font as Fnt
    for alert in alerts:
        cell = ws.cell(row=row, column=1, value=alert)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=8)
        row += 1

    _auto_width(ws, max_col_width=40)


def _build_recommendations(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("盘前推荐")
    headers = [
        "推荐日期", "模式", "排名", "股票代码", "股票名称",
        "主题名称", "总分", "人气分", "技术分", "空间分", "风险分",
        "主题强度", "主题加分", "推荐时收盘价", "大盘情绪分",
    ]
    df_sorted = df.sort_values(
        ["report_date", "mode", "rank"], ascending=[False, True, True]
    )
    rows = []
    for _, r in df_sorted.iterrows():
        rows.append({
            "推荐日期":     _date_fmt(r.get("report_date", "")),
            "模式":         _mode_cn(r.get("mode", "")),
            "排名":         _v(r.get("rank", "")) or "未记录",
            "股票代码":     _v(r.get("stock_code", "")) or "未记录",
            "股票名称":     _v(r.get("stock_name", "")) or "未记录",
            "主题名称":     _v(r.get("theme_name", "")) or "—",
            "总分":         _f(r.get("total_score", "")),
            "人气分":       _f(r.get("popularity_score", "")),
            "技术分":       _f(r.get("technical_score", "")),
            "空间分":       _f(r.get("space_score", "")),
            "风险分":       _f(r.get("risk_score", "")),
            "主题强度":     _f(r.get("theme_strength", "")),
            "主题加分":     _f(r.get("theme_bonus", "")),
            "推荐时收盘价": _f(r.get("recommended_close_price", "")),
            "大盘情绪分":   _f(r.get("market_sentiment", "")),
        })
    _write_table(ws, headers, rows, set(), 1)
    ws.freeze_panes = "A2"
    _auto_width(ws)


_SEC_REASON_CN = {
    "passed":                          "二次观察通过",
    "second_check_below_open":         "10:00 低于开盘价",
    "second_check_below_ma5":          "10:00 低于5日均线",
    "second_check_not_above_0935":     "10:00 未高于 9:36 价",
    "second_check_unable_limit_up":    "一字涨停买不进",
    "realtime_data_missing":           "实时行情获取失败",
    "realtime_price_invalid":          "价格数据无效",
}


def _sec_reason_cn(raw: str) -> str:
    s = _v(raw)
    if not s:
        return ""
    return "；".join(_SEC_REASON_CN.get(p.strip(), p.strip())
                     for p in s.split(";") if p.strip())


def _build_buy_check(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("买入确认(9点36)")
    headers = [
        "推荐日期", "模式", "股票代码", "股票名称", "主题名称",
        "是否模拟买入", "开盘涨幅%", "9点36价格",
        "模拟买入价", "滑点后买入价", "止损价",
        "买入原因", "不买原因", "备注",
        # —— V1.4 二次确认观察（仅观察，不参与正式收益）——
        "二次观察是否通过", "10点价格", "二次观察价", "二次观察结果", "二次观察时间",
    ]
    pct_headers = {"开盘涨幅%"}
    df_sorted = df.sort_values(
        ["report_date", "mode", "rank"], ascending=[False, True, True]
    )
    rows = []
    for _, r in df_sorted.iterrows():
        buy_sig  = _gb(r.get("buy_signal_0935", ""))
        unable   = _gb(r.get("unable_to_buy", ""))
        open_pct = _f(r.get("open_change_pct", ""))

        if buy_sig is True and unable is not True:
            buy_reason    = "买入条件全部满足"
            no_buy_reason = "—"
            remark        = ""
        elif unable is True:
            buy_reason    = "—"
            no_buy_reason = _v(r.get("unable_to_buy_reason", "")) or "涨停板"
            remark        = "满足买入条件但无法成交"
        elif _empty(r.get("open_price", "")):
            buy_reason    = "—"
            no_buy_reason = "9点36检查尚未运行"
            remark        = ""
        else:
            buy_reason    = "—"
            no_buy_reason = _notes_cn(r.get("notes", "")) or "未满足买入条件"
            remark        = ""

        # —— 二次确认观察字段（V1.4 实验性观察项）——
        sec_passed_raw = _v(r.get("second_check_passed", ""))
        sec_time       = _v(r.get("second_check_time", ""))
        if sec_time:
            sec_passed_cn = _bool_cn(sec_passed_raw)
            sec_reason_cn = _sec_reason_cn(r.get("second_check_reason", ""))
        else:
            sec_passed_cn = "—"
            sec_reason_cn = "—"

        rows.append({
            "推荐日期":     _date_fmt(r.get("report_date", "")),
            "模式":         _mode_cn(r.get("mode", "")),
            "股票代码":     _v(r.get("stock_code", "")),
            "股票名称":     _v(r.get("stock_name", "")),
            "主题名称":     _v(r.get("theme_name", "")) or "—",
            "是否模拟买入": _bool_cn(r.get("buy_signal_0935", "")),
            # open_change_pct stored as pct value (1.07 = 1.07%), /100 for Excel pct format
            "开盘涨幅%":    open_pct / 100 if open_pct is not None else None,
            "9点36价格":    _f(r.get("price_0935", "")),
            "模拟买入价":   _f(r.get("buy_price", "")),
            "滑点后买入价": _f(r.get("adjusted_buy_price", "")),
            "止损价":       _f(r.get("stop_price", "")),
            "买入原因":     buy_reason,
            "不买原因":     no_buy_reason,
            "备注":         remark or "—",
            "二次观察是否通过": sec_passed_cn,
            "10点价格":         _f(r.get("price_1000", "")),
            "二次观察价":       _f(r.get("second_check_observe_price", "")),
            "二次观察结果":     sec_reason_cn or "—",
            "二次观察时间":     sec_time or "—",
        })
    _write_table(ws, headers, rows, pct_headers, 1)
    ws.freeze_panes = "A2"
    _auto_width(ws)


def _build_post_buy(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("买入后复盘")
    headers = [
        "推荐日期", "模式", "股票代码", "股票名称",
        "模拟买入价", "滑点后买入价", "止损价",
        "次日开盘价", "次日最高价", "次日最低价", "次日收盘价",
        "次日最高收益", "最大回撤",
        "是否触发止损", "模拟卖出价", "模拟交易收益",
        "是否冲高3%", "是否冲高5%", "收盘是否盈利",
        "风险调整后是否成功", "路径是否不确定",
    ]
    # These are stored as decimals (0.03 = 3%) → Excel pct format
    pct_headers = {"次日最高收益", "最大回撤", "模拟交易收益"}

    df_bought = df[df["buy_signal_0935"].apply(lambda v: _gb(v) is True)]
    df_sorted = df_bought.sort_values(
        ["report_date", "mode", "rank"], ascending=[False, True, True]
    )
    rows = []
    for _, r in df_sorted.iterrows():
        rows.append({
            "推荐日期":         _date_fmt(r.get("report_date", "")),
            "模式":             _mode_cn(r.get("mode", "")),
            "股票代码":         _v(r.get("stock_code", "")),
            "股票名称":         _v(r.get("stock_name", "")),
            "模拟买入价":       _f(r.get("buy_price", "")),
            "滑点后买入价":     _f(r.get("adjusted_buy_price", "")),
            "止损价":           _f(r.get("stop_price", "")),
            "次日开盘价":       _f(r.get("t1_open", "")),
            "次日最高价":       _f(r.get("t1_high", "")),
            "次日最低价":       _f(r.get("t1_low", "")),
            "次日收盘价":       _f(r.get("t1_close", "")),
            "次日最高收益":     _f(r.get("t1_max_return", "")),
            "最大回撤":         _f(r.get("max_drawdown", "")),
            "是否触发止损":     _bool_cn(r.get("stop_loss_triggered", "")),
            "模拟卖出价":       _f(r.get("simulated_sell_price", "")),
            "模拟交易收益":     _f(r.get("simulated_trade_return", "")),
            "是否冲高3%":       _bool_cn(r.get("is_active_success", "")),
            "是否冲高5%":       _bool_cn(r.get("is_strong_surge", "")),
            "收盘是否盈利":     _bool_cn(r.get("is_close_success", "")),
            "风险调整后是否成功": _bool_cn(r.get("risk_adjusted_success", "")),
            "路径是否不确定":   _bool_cn(r.get("ambiguous_path", "")),
        })
    _write_table(ws, headers, rows, pct_headers, 1)
    ws.freeze_panes = "A2"
    _auto_width(ws)


def _build_not_bought(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("未买入跟踪")
    headers = [
        "推荐日期", "模式", "股票代码", "股票名称", "主题名称",
        "不买原因", "次日最高收益*", "次日收盘收益*", "是否错过大涨", "备注",
        # —— V1.4 二次确认观察 ——
        "二次观察是否通过", "10点价格", "二次观察价", "二次观察结果",
    ]
    pct_headers = {"次日最高收益*", "次日收盘收益*"}

    df_nb = df[~(df["buy_signal_0935"].apply(lambda v: _gb(v) is True) &
                 ~df["unable_to_buy"].apply(lambda v: _gb(v) is True))]
    df_sorted = df_nb.sort_values(
        ["report_date", "mode", "rank"], ascending=[False, True, True]
    )
    rows = []
    for _, r in df_sorted.iterrows():
        unable = _gb(r.get("unable_to_buy", ""))
        if unable is True:
            reason = _v(r.get("unable_to_buy_reason", "")) or "涨停板无法成交"
        elif _empty(r.get("open_price", "")):
            reason = "9:36检查尚未运行"
        else:
            reason = _notes_cn(r.get("notes", "")) or "未满足买入条件"

        ref = _f(r.get("recommended_close_price", ""))
        t1h = _f(r.get("t1_high", ""))
        t1c = _f(r.get("t1_close", ""))

        if ref and t1h:
            max_r   = (t1h - ref) / ref
            close_r = (t1c - ref) / ref if t1c else None
            missed  = "是" if max_r >= 0.03 else "否"
            remark  = "以推荐收盘价为参考，不计入正式统计"
        else:
            max_r = close_r = None
            missed = "未记录"
            remark = ""

        # —— 二次确认观察字段 ——
        sec_time = _v(r.get("second_check_time", ""))
        if sec_time:
            sec_passed_cn = _bool_cn(r.get("second_check_passed", ""))
            sec_reason_cn = _sec_reason_cn(r.get("second_check_reason", ""))
        else:
            sec_passed_cn = "—"
            sec_reason_cn = "—"

        rows.append({
            "推荐日期":      _date_fmt(r.get("report_date", "")),
            "模式":          _mode_cn(r.get("mode", "")),
            "股票代码":      _v(r.get("stock_code", "")),
            "股票名称":      _v(r.get("stock_name", "")),
            "主题名称":      _v(r.get("theme_name", "")) or "—",
            "不买原因":      reason,
            "次日最高收益*": max_r,
            "次日收盘收益*": close_r,
            "是否错过大涨":  missed,
            "备注":          remark or "—",
            "二次观察是否通过": sec_passed_cn,
            "10点价格":         _f(r.get("price_1000", "")),
            "二次观察价":       _f(r.get("second_check_observe_price", "")),
            "二次观察结果":     sec_reason_cn or "—",
        })
    _write_table(ws, headers, rows, pct_headers, 1)
    ws.freeze_panes = "A2"
    _auto_width(ws)


def _build_stats(wb, df: pd.DataFrame) -> None:
    ws = wb.create_sheet("模式统计")
    from openpyxl.styles import Font, Alignment

    intro = ws.cell(
        row=1, column=1,
        value="这个 sheet 用来比较 全A模式 和 主题龙头模式 哪个更有效。统计均基于 output/trade_review.csv 全量历史数据。",
    )
    intro.font      = Font(italic=True, color="595959", size=9)
    intro.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=4)
    ws.row_dimensions[1].height = 30

    df_full  = df[df["mode"].apply(_v) == "full"]
    df_theme = df[df["mode"].apply(_v) == "theme_auto"]
    sf       = _compute_stats(df_full)
    st       = _compute_stats(df_theme)

    def _r(v) -> str:
        return "N/A" if v is None else f"{v:.2f}"

    stat_rows = [
        {"指标": "推荐总数",          "说明": "历史所有推荐记录数",                     "全A模式": sf["total"],         "主题龙头模式": st["total"]},
        {"指标": "有效样本数",         "说明": "已完成9:36检查的推荐数",                  "全A模式": sf["n_valid"],        "主题龙头模式": st["n_valid"]},
        {"指标": "买入触发数",         "说明": "满足9:36买入条件的推荐数",               "全A模式": sf["n_triggered"],    "主题龙头模式": st["n_triggered"]},
        {"指标": "买入触发率",         "说明": "触发数 / 有效样本数",                    "全A模式": _pct_str(sf["bsr"]),  "主题龙头模式": _pct_str(st["bsr"])},
        {"指标": "含T+1数据的交易数",  "说明": "有完整T+1数据可以复盘的交易数",          "全A模式": sf["n_traded"],       "主题龙头模式": st["n_traded"]},
        {"指标": "风险调整后成功率",   "说明": "冲高≥3% 且未先触-3%止损线的比例",       "全A模式": _pct_str(sf["risk_rate"]),   "主题龙头模式": _pct_str(st["risk_rate"])},
        {"指标": "冲高3%比例",         "说明": "次日最高价比买入价高出3%以上的概率",     "全A模式": _pct_str(sf["active_rate"]), "主题龙头模式": _pct_str(st["active_rate"])},
        {"指标": "冲高5%比例",         "说明": "次日最高价比买入价高出5%以上的概率",     "全A模式": _pct_str(sf["surge_rate"]),  "主题龙头模式": _pct_str(st["surge_rate"])},
        {"指标": "收盘胜率",           "说明": "次日收盘价高于买入价的概率",             "全A模式": _pct_str(sf["close_rate"]),  "主题龙头模式": _pct_str(st["close_rate"])},
        {"指标": "止损率",             "说明": "触发-3%止损线的概率",                   "全A模式": _pct_str(sf["stop_rate"]),   "主题龙头模式": _pct_str(st["stop_rate"])},
        {"指标": "平均盈利",           "说明": "盈利交易的平均收益率",                  "全A模式": _gain_str(sf["avg_gain"]),   "主题龙头模式": _gain_str(st["avg_gain"])},
        {"指标": "平均亏损",           "说明": "亏损交易的平均亏损率",                  "全A模式": _gain_str(sf["avg_loss"]),   "主题龙头模式": _gain_str(st["avg_loss"])},
        {"指标": "盈亏比",             "说明": "平均盈利÷平均亏损，>1说明赚多亏少",    "全A模式": _r(sf["wl_ratio"]),          "主题龙头模式": _r(st["wl_ratio"])},
    ]

    _write_table(ws, ["指标", "说明", "全A模式", "主题龙头模式"], stat_rows, set(), start_row=2)
    ws.freeze_panes = "A3"
    _auto_width(ws, max_col_width=60)


def _build_period_review(wb, df: pd.DataFrame, period_type: str) -> None:
    """
    Build 周复盘 / 月复盘 sheet —— V1.4 改版后含 7 个区块：
      1 总览  2 推荐明细  3 模拟买入明细  4 未买入明细
      5 二次确认观察  6 不买原因统计  7 本周/本月结论
    数据通过 periodic_review.{_run_review的 detail builders} 计算，保证 MD 与 Excel 一致。
    """
    # 复用 periodic_review 里所有 detail builders + 标题逻辑
    import periodic_review as pr

    today = _date.today()
    if period_type == "weekly":
        start_str, end_str, label, period_title = pr._current_week_range()
        sheet_name = "周复盘"
        period_cn  = "周"
    else:
        start_str, end_str, label, period_title = pr._current_month_range()
        sheet_name = "月复盘"
        period_cn  = "月"

    ws = wb.create_sheet(sheet_name)
    from openpyxl.styles import Font, Alignment

    # —— 标题行 ——
    title_val = f"{period_title}"
    tc = ws.cell(row=1, column=1, value=title_val)
    tc.font      = Font(bold=True, size=12, color=_C_TITLE)
    tc.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=14)
    ws.row_dimensions[1].height = 22

    # —— 筛本期数据 ——
    mask      = (df["report_date"].astype(str) >= start_str) & (df["report_date"].astype(str) <= end_str)
    df_period = df[mask]

    if df_period.empty:
        ws.cell(row=2, column=1,
                value=f"本{period_cn}（{_date_fmt(start_str)} 至 {_date_fmt(end_str)}）暂无推荐数据。")
        _auto_width(ws, max_col_width=60)
        return

    # —— 统计指标 ——
    df_full  = df_period[df_period["mode"].apply(_v) == "full"]
    df_theme = df_period[df_period["mode"].apply(_v) == "theme_auto"]
    overall  = _compute_stats(df_period)
    sf       = _compute_stats(df_full)
    st       = _compute_stats(df_theme)

    # —— 明细行 ——
    rec_rows         = pr._build_recommendations_detail(df_period)
    bought_rows      = pr._build_bought_detail(df_period)
    not_bought_rows  = pr._build_not_bought_detail(df_period)
    second_check_rows = pr._build_second_check_detail(df_period)
    no_buy_with_names = pr._count_no_buy_reasons_with_names(df_period)
    plain_summary    = pr._generate_plain_summary(
        period_cn, overall, sf, st,
        no_buy_with_names, bought_rows, second_check_rows,
    )

    has_traded = overall["n_traded"] > 0
    na_traded  = "暂无已完成T+1样本"

    def _pp(v): return _pct_str(v) if v is not None else (na_traded if not has_traded else "N/A")
    def _rr(v): return f"{v:.2f}" if v is not None else (na_traded if not has_traded else "N/A")
    def _gg(v): return _gain_str(v) if v is not None else (na_traded if not has_traded else "N/A")
    def _pp_bsr(v): return _pct_str(v) if v is not None else "（无9:36样本）"

    # ─── 区块 1：本周/月总览 ──────────────────────────────────────
    row = 2
    row = _write_section_label(ws, row, f"1. 本{period_cn}总览", ncols=14)
    stat_rows = [
        {"指标": "推荐总数",            "全部": overall["total"],          "全A模式": sf["total"],          "主题龙头模式": st["total"]},
        {"指标": "完成9:36检查",        "全部": overall["n_valid"],        "全A模式": sf["n_valid"],        "主题龙头模式": st["n_valid"]},
        {"指标": "触发模拟买入",        "全部": overall["n_triggered"],    "全A模式": sf["n_triggered"],    "主题龙头模式": st["n_triggered"]},
        {"指标": "买入触发率",          "全部": _pp_bsr(overall["bsr"]),   "全A模式": _pp_bsr(sf["bsr"]),   "主题龙头模式": _pp_bsr(st["bsr"])},
        {"指标": "已完成T+1复盘",       "全部": overall["n_traded"],       "全A模式": sf["n_traded"],       "主题龙头模式": st["n_traded"]},
        {"指标": "风险调整后成功率",     "全部": _pp(overall.get("risk_rate")),   "全A模式": _pp(sf.get("risk_rate")),   "主题龙头模式": _pp(st.get("risk_rate"))},
        {"指标": "冲高3%比例",          "全部": _pp(overall.get("active_rate")), "全A模式": _pp(sf.get("active_rate")), "主题龙头模式": _pp(st.get("active_rate"))},
        {"指标": "冲高5%比例",          "全部": _pp(overall.get("surge_rate")),  "全A模式": _pp(sf.get("surge_rate")),  "主题龙头模式": _pp(st.get("surge_rate"))},
        {"指标": "收盘胜率",            "全部": _pp(overall.get("close_rate")),  "全A模式": _pp(sf.get("close_rate")),  "主题龙头模式": _pp(st.get("close_rate"))},
        {"指标": "止损率",              "全部": _pp(overall.get("stop_rate")),   "全A模式": _pp(sf.get("stop_rate")),   "主题龙头模式": _pp(st.get("stop_rate"))},
        {"指标": "平均盈利",            "全部": _gg(overall.get("avg_gain")),    "全A模式": _gg(sf.get("avg_gain")),    "主题龙头模式": _gg(st.get("avg_gain"))},
        {"指标": "平均亏损",            "全部": _gg(overall.get("avg_loss")),    "全A模式": _gg(sf.get("avg_loss")),    "主题龙头模式": _gg(st.get("avg_loss"))},
        {"指标": "盈亏比",              "全部": _rr(overall.get("wl_ratio")),    "全A模式": _rr(sf.get("wl_ratio")),    "主题龙头模式": _rr(st.get("wl_ratio"))},
    ]
    row = _write_table(ws, ["指标", "全部", "全A模式", "主题龙头模式"], stat_rows, set(), start_row=row)

    if not has_traded:
        cell = ws.cell(row=row, column=1,
                       value=f"⚠️ 本{period_cn}暂无已完成 T+1 复盘样本，收益类指标暂不可用（等待 T+1 数据补全后自动重算）。")
        cell.font      = Font(italic=True, color="C00000", size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
        row += 1
    row += 1  # blank

    # ─── 区块 2：推荐明细 ─────────────────────────────────────────
    row = _write_section_label(ws, row, f"2. 本{period_cn}推荐明细", ncols=14)
    if rec_rows:
        headers = [
            "推荐日期", "模式", "排名", "代码", "名称", "主题",
            "总分", "人气", "技术", "主题强度",
            "9:36已查", "模拟买入", "买入价", "买入/不买原因", "当前状态",
        ]
        excel_rows = []
        for r in rec_rows:
            excel_rows.append({
                "推荐日期":     _date_fmt(r["report_date"]),
                "模式":         _mode_cn(r["mode"]),
                "排名":         r["rank"],
                "代码":         r["code"],
                "名称":         r["name"],
                "主题":         r["theme_name"],
                "总分":         r["total_score"],
                "人气":         r["popularity"],
                "技术":         r["technical"],
                "主题强度":     r["theme_strength"] if r["theme_strength"] is not None else "—",
                "9:36已查":     r["checked_0935"],
                "模拟买入":     r["buy_signal"],
                "买入价":       r["buy_price"]   if r["buy_price"]   is not None else "—",
                "买入/不买原因": r["reason_text"],
                "当前状态":     r["status"],
            })
        row = _write_table(ws, headers, excel_rows, set(), start_row=row)
    else:
        ws.cell(row=row, column=1, value=f"本{period_cn}无推荐数据。")
        row += 1
    row += 1

    # ─── 区块 3：模拟买入明细 ─────────────────────────────────────
    row = _write_section_label(ws, row, f"3. 本{period_cn}模拟买入明细", ncols=14)
    if bought_rows:
        headers = [
            "推荐日期", "模式", "代码", "名称", "主题",
            "买入价", "滑点后买入价", "止损价", "买入原因",
            "T+1状态", "次日最高收益", "最大回撤", "模拟收益",
            "是否止损", "风险调整后成功",
        ]
        pct_headers = {"次日最高收益", "最大回撤", "模拟收益"}
        excel_rows = []
        for b in bought_rows:
            t1_status = "已完成" if b["t1_done"] else "等待T+1复盘"
            t1_max = b["t1_max_return"]   if b["t1_done"] else "等待T+1复盘"
            t1_dd  = b["max_drawdown"]    if b["t1_done"] else "等待T+1复盘"
            t1_ret = b["trade_return"]    if b["t1_done"] else "等待T+1复盘"
            excel_rows.append({
                "推荐日期":      _date_fmt(b["report_date"]),
                "模式":          _mode_cn(b["mode"]),
                "代码":          b["code"],
                "名称":          b["name"],
                "主题":          b["theme_name"],
                "买入价":        b["buy_price"],
                "滑点后买入价":  b["adj_buy_price"],
                "止损价":        b["stop_price"],
                "买入原因":      b["buy_reason"],
                "T+1状态":       t1_status,
                "次日最高收益":  t1_max,
                "最大回撤":      t1_dd,
                "模拟收益":      t1_ret,
                "是否止损":      b["stop_triggered"],
                "风险调整后成功": b["risk_success"],
            })
        row = _write_table(ws, headers, excel_rows, pct_headers, start_row=row)
    else:
        ws.cell(row=row, column=1, value=f"本{period_cn}未触发任何模拟买入。")
        row += 1
    row += 1

    # ─── 区块 4：未买入明细 ───────────────────────────────────────
    row = _write_section_label(ws, row, f"4. 本{period_cn}未买入明细", ncols=14)
    if not_bought_rows:
        headers = [
            "推荐日期", "模式", "代码", "名称", "主题",
            "不买主因", "辅助原因", "9:36价", "开盘价", "5日线",
            "二次确认观察", "是否错过大涨", "T+1最高收益", "T+1收盘收益",
        ]
        pct_headers = {"T+1最高收益", "T+1收盘收益"}
        excel_rows = []
        for n in not_bought_rows:
            excel_rows.append({
                "推荐日期":     _date_fmt(n["report_date"]),
                "模式":         _mode_cn(n["mode"]),
                "代码":         n["code"],
                "名称":         n["name"],
                "主题":         n["theme_name"],
                "不买主因":     n["hard_reasons"],
                "辅助原因":     n["soft_reasons"],
                "9:36价":       n["price_0935"]    if n["price_0935"] is not None else "—",
                "开盘价":       n["open_price"]    if n["open_price"] is not None else "—",
                "5日线":        n["ma5"]           if n["ma5"]        is not None else "—",
                "二次确认观察": n["sec_check"],
                "是否错过大涨": n["missed_big"],
                "T+1最高收益":  n["t1_max_return"]    if n["t1_max_return"]    is not None else "等待T+1观察",
                "T+1收盘收益":  n["t1_close_return"]  if n["t1_close_return"]  is not None else "等待T+1观察",
            })
        row = _write_table(ws, headers, excel_rows, pct_headers, start_row=row)
    else:
        ws.cell(row=row, column=1, value=f"本{period_cn}所有推荐均已触发模拟买入。")
        row += 1
    row += 1

    # ─── 区块 5：二次确认观察 ─────────────────────────────────────
    row = _write_section_label(ws, row,
        f"5. 本{period_cn}二次确认观察 — 仅记录不买入，不计入正式收益", ncols=14)
    if second_check_rows:
        headers = [
            "推荐日期", "模式", "代码", "名称",
            "9:36 不买原因", "10:00价", "二次确认是否通过", "二次确认原因",
            "观察价", "后续T+1表现",
        ]
        excel_rows = []
        for s in second_check_rows:
            excel_rows.append({
                "推荐日期":         _date_fmt(s["report_date"]),
                "模式":             _mode_cn(s["mode"]),
                "代码":             s["code"],
                "名称":             s["name"],
                "9:36 不买原因":    s["orig_reason"],
                "10:00价":          s["price_1000"]    if s["price_1000"]    is not None else "—",
                "二次确认是否通过": s["sec_passed"],
                "二次确认原因":     s["sec_reason"],
                "观察价":           s["observe_price"] if s["observe_price"] is not None else "—",
                "后续T+1表现":      s["t1_followup"],
            })
        row = _write_table(ws, headers, excel_rows, set(), start_row=row)
    else:
        ws.cell(row=row, column=1,
                value=f"本{period_cn}未做过 10:00 二次确认观察（或当日 9:36 全部已买入/无可观察样本）。")
        row += 1
    row += 1

    # ─── 区块 6：不买原因统计 ─────────────────────────────────────
    row = _write_section_label(ws, row, f"6. 本{period_cn}不买原因统计", ncols=14)
    if no_buy_with_names:
        headers = ["不买原因", "次数", "涉及股票"]
        excel_rows = []
        for r in no_buy_with_names:
            excel_rows.append({
                "不买原因":  r["reason_cn"],
                "次数":      r["count"],
                "涉及股票":  "、".join(r["stocks"]) if r["stocks"] else "—",
            })
        row = _write_table(ws, headers, excel_rows, set(), start_row=row)
    else:
        ws.cell(row=row, column=1, value=f"本{period_cn}无未买入记录。")
        row += 1
    row += 1

    # ─── 区块 7：本周/月结论 ──────────────────────────────────────
    row = _write_section_label(ws, row, f"7. 本{period_cn}结论", ncols=14)
    for s in plain_summary:
        cell = ws.cell(row=row, column=1, value=f"• {s}")
        cell.font      = Font(size=10)
        cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=14)
        row += 1

    ws.freeze_panes = "A3"
    _auto_width(ws, max_col_width=60)


def _date_fmt(s: str) -> str:
    """Format YYYYMMDD → YYYY-MM-DD, pass through otherwise."""
    if len(s) == 8 and s.isdigit():
        return f"{s[:4]}-{s[4:6]}-{s[6:8]}"
    return s


def _build_field_notes(wb) -> None:
    ws = wb.create_sheet("字段说明")
    from openpyxl.styles import Font, Alignment

    intro = ws.cell(row=1, column=1,
                    value="字段含义说明 — 用大白话解释各指标的意思。完整说明见 output/字段说明.md")
    intro.font      = Font(italic=True, color="595959", size=9)
    intro.alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=2)

    (hdr_fill, hdr_font, _, _, alt_fill,
     _, _, _, wrap_left, center_va, _) = _styles()

    # Header row
    for col_idx, h in enumerate(["字段名", "含义说明"], 1):
        cell = ws.cell(row=2, column=col_idx, value=h)
        cell.font      = hdr_font
        cell.fill      = hdr_fill
        cell.alignment = center_va

    row_idx  = 3
    data_row = 0  # counts non-section rows for alternating colour
    for name, desc in _FIELD_NOTES:
        if name == "":
            # Section header — spans both columns
            _write_section_label(ws, row_idx, desc, ncols=2)
            row_idx += 1
        else:
            data_row += 1
            use_alt = data_row % 2 == 0
            name_cell = ws.cell(row=row_idx, column=1, value=name)
            desc_cell = ws.cell(row=row_idx, column=2, value=desc)
            name_cell.alignment = Alignment(horizontal="left", vertical="center")
            desc_cell.alignment = Alignment(horizontal="left", vertical="center",
                                            wrap_text=True)
            if use_alt:
                name_cell.fill = alt_fill
                desc_cell.fill = alt_fill
            row_idx += 1

    ws.freeze_panes = "A3"
    _auto_width(ws, max_col_width=70)


# ── public entry point ────────────────────────────────────────────────────────

def generate_excel_report() -> None:
    """生成 Excel 总表。失败静默不中断主流程。"""
    try:
        _generate()
        logger.info(f"[excel_report] Excel已更新: {EXCEL_PATH.name}")
    except Exception as e:
        logger.warning(f"[excel_report] Excel生成失败（不影响主流程）: {e}")


def _generate() -> None:
    import openpyxl  # noqa: F401 — verifies installation early

    OUTPUT_DIR.mkdir(exist_ok=True)
    from openpyxl import Workbook
    wb = Workbook()
    wb.remove(wb.active)  # remove default "Sheet"

    if not CSV_PATH.exists():
        ws = wb.create_sheet("今日总览")
        ws.cell(row=1, column=1, value="暂无数据，请先运行 run.py")
        wb.save(EXCEL_PATH)
        return

    df = pd.read_csv(CSV_PATH, dtype=str, keep_default_na=False, encoding="utf-8-sig")

    _build_overview(wb, df)
    _build_recommendations(wb, df)
    _build_buy_check(wb, df)
    _build_post_buy(wb, df)
    _build_not_bought(wb, df)
    _build_stats(wb, df)
    _build_period_review(wb, df, "weekly")
    _build_period_review(wb, df, "monthly")
    _build_field_notes(wb)

    wb.save(EXCEL_PATH)
